import openpyxl
import numpy as np

from em1234567 import EM1234567

def getFundListFromXlsx(f_xlsx, f_save_dir):
	wb = openpyxl.load_workbook(f_xlsx)
	sh = wb['INFO']
	for i in range(1, sh.max_row+1):
		fund_code = sh.cell(row=i, column=1).value
		em = EM1234567(fund_code)
		try:
			xy = em.getHistoryPlotJson()
			np.savetxt('%s/%s.csv' % (f_save_dir, fund_code), np.array(xy), delimiter=',')
		except:
			print('Unable to find [%s].' % fund_code)

def getFundListFromCode(fund_code, f_save_dir):
	em = EM1234567(fund_code)
	try:
		xy = em.getHistoryPlotJson()
		np.savetxt('%s/%s.csv' % (f_save_dir, fund_code), np.array(xy), delimiter=',')
	except:
		print('Unable to find [%s].' % fund_code)


if __name__ == '__main__':
	import sys
	print('python3 local-update.py <-f <xxx.xlsx>|-c <code>> <dir>')
	if sys.argv[1] == '-f':
		getFundListFromXlsx(sys.argv[2], sys.argv[3])
	elif sys.argv[1] == '-c':
		getFundListFromCode(sys.argv[2], sys.argv[3])