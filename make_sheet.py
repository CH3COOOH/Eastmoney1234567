from em1234567 import EM1234567
import openpyxl

def updateSheet(fname, isHistory):
	wb = openpyxl.load_workbook(fname)
	sh = wb['INFO']
	for i in range(1, sh.max_row):
		fund_code = sh.cell(row=i+1, column=1).value
		em = EM1234567(fund_code)
		try:
			realtime = em.getRealtimeInfo()
			print('Updating [%s] %s...' % (fund_code, realtime['name']))
			sh.cell(row=i+1, column=2).value = realtime['name']
			sh.cell(row=i+1, column=3).value = float(realtime['gszzl'])
			sh.cell(row=i+1, column=4).value = realtime['gztime']
			if isHistory == '1':
				history = em.getHistoryRate()
				sh.cell(row=i+1, column=5).value = history['1m']
				sh.cell(row=i+1, column=6).value = history['3m']
				sh.cell(row=i+1, column=7).value = history['6m']
				sh.cell(row=i+1, column=8).value = history['1y']
		except:
			print('Unable to find [%s].' % fund_code)
	wb.save(fname)

if __name__ == '__main__':
	import sys
	updateSheet(sys.argv[1], sys.argv[2])
	
